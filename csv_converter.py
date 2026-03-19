"""
csv_converter.py
----------------
Grafische App zum Konvertieren von Sekunden-CSV zu Minuten-CSV oder Excel.
Benötigte Pakete: pip install customtkinter pandas openpyxl

Zu einer .app/.exe kompilieren:
  pip install pyinstaller
  pyinstaller --onefile --windowed csv_converter.py
"""

import customtkinter as ctk
from tkinter import filedialog
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import threading
import os


# ── Design ────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

FONT_TITLE  = ("Georgia", 22, "bold")
FONT_LABEL  = ("Georgia", 13)
FONT_SMALL  = ("Georgia", 11)
FONT_BUTTON = ("Georgia", 13, "bold")

BG     = "#0f1117"
CARD   = "#1a1d27"
ACCENT = "#4f8ef7"
ACCENT2= "#a78bfa"
SUCCESS= "#34d399"
ERROR  = "#f87171"
TEXT   = "#e8eaf0"
MUTED  = "#6b7280"


# ── Hilfsfunktionen ───────────────────────────────────────────────────────────
def load_csv(input_file):
    """CSV einlesen und datetime-Spalte erzeugen."""
    df = pd.read_csv(input_file, encoding="utf-8-sig", decimal=",")
    time_str = df["Time"].str.replace(r"\s*\+\d{2}:\d{2}$", "", regex=True).str.strip()
    combined = df["Date"].str.strip() + " " + time_str
    # Format automatisch erkennen: wenn Tag-Teil > 12 → mm.dd.yyyy, sonst dd.mm.yyyy
    sample_parts = df["Date"].iloc[0].split(".")
    is_american = int(sample_parts[1]) > 12
    fmt = "%m.%d.%Y %H:%M:%S.%f" if is_american else "%d.%m.%Y %H:%M:%S.%f"
    df["datetime"] = pd.to_datetime(combined, format=fmt, errors="coerce")
    return df


def aggregate_minutes(df, decimals):
    """Sekundendaten zu Minutenmittelwerten aggregieren."""
    df["minute"] = df["datetime"].dt.floor("min")
    exclude = {"Date", "Time", "Diff", "Sequence", "datetime", "minute"}
    numeric_cols = [
        c for c in df.columns
        if c not in exclude and pd.api.types.is_numeric_dtype(df[c])
    ]
    df_out = (
        df.groupby("minute")[numeric_cols]
        .mean(numeric_only=True)
        .reset_index()
    )
    if decimals == 0:
        df_out[numeric_cols] = df_out[numeric_cols].round(0).astype("Int64")
    else:
        df_out[numeric_cols] = df_out[numeric_cols].round(decimals)
    df_out.rename(columns={"minute": "Timestamp"}, inplace=True)
    return df_out


def style_excel_sheet(ws, df):
    """Kopfzeile einfärben, Zahlen formatieren, Spaltenbreite anpassen."""
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Zahlen als Zahl formatieren
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.##"
            elif isinstance(cell.value, int):
                cell.number_format = "0"

    # Spaltenbreite basierend auf Inhalt
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max()
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Kopfzeile fixieren
    ws.freeze_panes = "A2"


def save_excel(path, sheets: dict):
    """Mehrere DataFrames als benannte Tabs in eine Excel-Datei speichern."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_excel_sheet(ws, df)


# ── Konvertierungs-Modi ────────────────────────────────────────────────────────
def convert_csv_to_excel(input_file, output_file, decimals):
    """Nur CSV → Excel (keine Zeitaggregation)."""
    df = load_csv(input_file)
    exclude = {"datetime"}
    df_clean = df.drop(columns=[c for c in exclude if c in df.columns])
    save_excel(output_file, {"Daten": df_clean})
    return len(df_clean), len(df_clean.columns)


def convert_csv_to_minutes_csv(input_file, output_file, decimals):
    """CSV → Minuten-CSV."""
    df = load_csv(input_file)
    df_out = aggregate_minutes(df, decimals)
    df_out.to_csv(output_file, index=False, decimal=".")
    return len(df_out), len(df_out.columns) - 1


def convert_csv_to_minutes_excel(input_file, output_file, decimals):
    """CSV → Excel mit zwei Tabs: Rohdaten + Minutendaten."""
    df = load_csv(input_file)
    df_raw = df.drop(columns=["datetime"], errors="ignore")
    df_min = aggregate_minutes(df, decimals)
    save_excel(output_file, {
        "Rohdaten":     df_raw,
        "Minutendaten": df_min
    })
    return len(df_min), len(df_min.columns) - 1


# ── App ───────────────────────────────────────────────────────────────────────
MODES = {
    "CSV → Excel":                  ("excel",   ".xlsx", convert_csv_to_excel),
    "CSV → Minuten-CSV":            ("csv",     ".csv",  convert_csv_to_minutes_csv),
    "CSV → Minuten-Excel":          ("excel",   ".xlsx", convert_csv_to_minutes_excel),
}

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("CSV Konverter")
        self.geometry("640x640")
        self.resizable(False, False)
        self.configure(fg_color=BG)

        self._input_path  = ctk.StringVar()
        self._output_path = ctk.StringVar()
        self._decimals_var = ctk.StringVar(value="2")
        self._mode        = ctk.StringVar(value=list(MODES.keys())[0])

        self._build_ui()

    def _build_ui(self):
        # Titel
        title_frame = ctk.CTkFrame(self, fg_color="transparent")
        title_frame.pack(pady=(32, 2), padx=48, anchor="w")
        ctk.CTkLabel(title_frame, text="CSV", font=("Georgia", 32, "bold"),
                     text_color=ACCENT).pack(side="left")
        ctk.CTkLabel(title_frame, text=" Konverter", font=("Georgia", 32, "bold"),
                     text_color=TEXT).pack(side="left")
        ctk.CTkLabel(self, text="Medizinische Monitordaten aufbereiten",
                     font=FONT_SMALL, text_color=MUTED).pack(anchor="w", padx=48, pady=(0, 20))

        # Modus-Auswahl
        card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=14)
        card.pack(fill="x", padx=48, pady=8)
        ctk.CTkLabel(card, text="Modus", font=FONT_LABEL,
                     text_color=TEXT).pack(anchor="w", padx=20, pady=(14, 6))
        seg = ctk.CTkSegmentedButton(
            card, values=list(MODES.keys()),
            variable=self._mode,
            font=FONT_SMALL,
            selected_color=ACCENT, selected_hover_color=ACCENT2,
            unselected_color="#2d3148", unselected_hover_color="#3d4168",
            command=self._on_mode_change
        )
        seg.pack(fill="x", padx=20, pady=(0, 14))

        # Eingabe-Datei
        self._file_row("Eingabe-Datei", "CSV mit Monitor-Daten",
                       self._input_path, "Datei wählen", self._pick_input)

        # Ausgabe-Datei
        self._file_row("Ausgabe-Datei", "Speicherort der neuen Datei",
                       self._output_path, "Speicherort", self._pick_output)

        # Nachkommastellen (nur bei Minuten-Modi sichtbar)
        self._dec_card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=14)
        self._dec_card.pack(fill="x", padx=48, pady=8)
        row = ctk.CTkFrame(self._dec_card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=14)
        ctk.CTkLabel(row, text="Nachkommastellen", font=FONT_LABEL,
                     text_color=TEXT).pack(side="left")
        ctk.CTkLabel(row, text="(0 – 6)", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=8)
        self._dec_entry = ctk.CTkEntry(
            row, textvariable=self._decimals_var,
            width=60, height=36, font=FONT_LABEL,
            fg_color="#0f1117", border_color="#2d3148",
            text_color=ACCENT, justify="center"
        )
        self._dec_entry.pack(side="right")

        # Start-Button
        self._btn = ctk.CTkButton(
            self, text="▶  Konvertieren", font=FONT_BUTTON, height=48,
            fg_color=ACCENT, hover_color=ACCENT2, corner_radius=12,
            command=self._start
        )
        self._btn.pack(fill="x", padx=48, pady=(16, 8))

        # Status + Progressbar
        self._status = ctk.CTkLabel(self, text="", font=FONT_SMALL,
                                     text_color=MUTED, wraplength=520)
        self._status.pack(pady=4)
        self._progress = ctk.CTkProgressBar(
            self, fg_color="#2d3148", progress_color=ACCENT, height=6)
        self._progress.set(0)
        self._progress.pack(fill="x", padx=48, pady=(4, 0))

        # Nachkommastellen ausblenden wenn Modus kein Minuten hat
        self._on_mode_change(self._mode.get())

    def _file_row(self, label, desc, var, btn_text, cmd):
        card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=14)
        card.pack(fill="x", padx=48, pady=8)
        info = ctk.CTkFrame(card, fg_color="transparent")
        info.pack(fill="x", padx=20, pady=(14, 4))
        ctk.CTkLabel(info, text=label, font=FONT_LABEL, text_color=TEXT).pack(side="left")
        ctk.CTkLabel(info, text=desc,  font=FONT_SMALL, text_color=MUTED).pack(side="left", padx=10)
        row = ctk.CTkFrame(card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 14))
        ctk.CTkEntry(row, textvariable=var, font=FONT_SMALL,
                     fg_color="#0f1117", border_color="#2d3148",
                     text_color=TEXT, placeholder_text="Pfad ...",
                     height=36).pack(side="left", fill="x", expand=True, padx=(0, 10))
        ctk.CTkButton(row, text=btn_text, font=FONT_SMALL, width=120, height=36,
                      fg_color="#2d3148", hover_color=ACCENT,
                      corner_radius=8, command=cmd).pack(side="right")

    def _on_mode_change(self, val):
        # Nachkommastellen nur bei Minuten-Modi anzeigen
        if "Minuten" in val:
            self._dec_card.pack(fill="x", padx=48, pady=8)
        else:
            self._dec_card.pack_forget()
        # Ausgabepfad-Extension aktualisieren
        self._update_output_extension()

    def _update_output_extension(self):
        out = self._output_path.get()
        if not out:
            return
        _, ext = MODES[self._mode.get()][0], MODES[self._mode.get()][1]
        base = os.path.splitext(out)[0]
        self._output_path.set(base + ext)

    def _update_dec_label(self, val=None):
        pass  # nicht mehr benötigt (Eingabefeld)

    def _pick_input(self):
        path = filedialog.askopenfilename(
            title="Eingabe-CSV wählen",
            filetypes=[("CSV Dateien", "*.csv *.CSV"), ("Alle Dateien", "*.*")]
        )
        if path:
            self._input_path.set(path)
            ext = MODES[self._mode.get()][1]
            base = os.path.splitext(path)[0]
            suffix = "_minuten" if "Minuten" in self._mode.get() else ""
            self._output_path.set(base + suffix + ext)

    def _pick_output(self):
        ext  = MODES[self._mode.get()][1]
        ft   = [("Excel", "*.xlsx")] if ext == ".xlsx" else [("CSV", "*.csv")]
        path = filedialog.asksaveasfilename(
            title="Ausgabe speichern",
            defaultextension=ext,
            filetypes=ft
        )
        if path:
            self._output_path.set(path)

    def _start(self):
        inp  = self._input_path.get().strip()
        out  = self._output_path.get().strip()
        try:
            decs = int(self._decimals_var.get())
            if not 0 <= decs <= 6:
                raise ValueError
        except ValueError:
            self._set_status("⚠  Nachkommastellen: bitte eine Zahl zwischen 0 und 6 eingeben.", ERROR)
            return

        if not inp:
            self._set_status("⚠  Bitte eine Eingabe-Datei wählen.", ERROR); return
        if not os.path.exists(inp):
            self._set_status("⚠  Eingabe-Datei nicht gefunden.", ERROR); return
        if not out:
            self._set_status("⚠  Bitte einen Speicherort wählen.", ERROR); return

        self._btn.configure(state="disabled", text="⏳  Läuft ...")
        self._progress.configure(mode="indeterminate")
        self._progress.start()
        self._set_status("Verarbeite ...", MUTED)
        threading.Thread(target=self._run, args=(inp, out, decs), daemon=True).start()

    def _run(self, inp, out, decs):
        try:
            fn = MODES[self._mode.get()][2]
            rows, cols = fn(inp, out, decs)
            self.after(0, self._done,
                       f"✓  Fertig!  {rows} Zeilen · {cols} Spalten\n→ {out}", SUCCESS)
        except Exception as e:
            self.after(0, self._done, f"✗  Fehler: {e}", ERROR)

    def _done(self, msg, color):
        self._progress.stop()
        self._progress.configure(mode="determinate")
        self._progress.set(1 if color == SUCCESS else 0)
        self._set_status(msg, color)
        self._btn.configure(state="normal", text="▶  Konvertieren")

    def _set_status(self, msg, color):
        self._status.configure(text=msg, text_color=color)


if __name__ == "__main__":
    App().mainloop()
